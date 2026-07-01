import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment


# SQLite 데이터베이스에 연결
db_path = os.path.join(os.getcwd(), "test.db")
conn = sqlite3.connect('C:/Users/ITis/Desktop/my_workspace/pikeview/pikeview/backend/test.db', check_same_thread=False)
cursor = conn.cursor()

cursor.execute("SELECT Datetime, Result, H_BootTear, H_BootCurl, H_BootBurr, H_BootAssembly, H_GreaseOut, \
                H_PBallDamage, H_PBallLot, H_CRingErr, H_CRingTwist, H_ORingErr, H_ORingTwist, SocketDamage, \
                SocketGroove, L_BootTear, L_BootCurl, L_BootBurr, L_BootAssembly, L_GreaseOut, L_PBallDamage, \
                L_CRingErr, L_CRingTwist, L_ORingErr, L_ORingTwist, DefectGroup \
                FROM ctrDB")
rows = cursor.fetchall()

fieldDic = {
    "ClientName": "생산라인",
    "Part": "부품",
    "Datetime": "폴더명",
    "Result": "검사결과",
    "H_BootTear": "1차 부트 찢어짐",
    "H_BootCurl": "1차 부트 말림",
    "H_BootBurr": "1차 부트 Burr 유/무",
    "H_BootAssembly": "1차 부트 조립 상태",
    "H_GreaseOut": "1차 목부 그리스 누유",
    "H_PBallDamage": "1차 P/Ball 찍힘 및 긁힘",
    "H_PBallLot": "1차 P/Ball 로트 타각 식별",
    "H_CRingErr": "1차 C/Ring 유/무",
    "H_CRingTwist": "1차 C/Ring 꼬임",
    "H_ORingErr": "1차 O/Ring 유/무",
    "H_ORingTwist": "1차 O/Ring 꼬임",
    "SocketDamage": "소켓 외경 찍힘 및 긁힘",
    "SocketGroove": "소켓 외경 홈 그루브 유무",
    "L_BootTear": "2차 부트 찢어짐",
    "L_BootCurl": "2차 부트 말림",
    "L_BootBurr": "2차 부트 Burr 유/무",
    "L_BootAssembly": "2차 부트 조립 상태",
    "L_GreaseOut": "2차 목부 그리스 누유",
    "L_PBallDamage": "2차 P/Ball 찍힘 및 긁힘",
    "L_CRingErr": "2차 C/Ring 유/무",
    "L_CRingTwist": "2차 C/Ring 꼬임",
    "L_ORingErr": "2차 O/Ring 유/무",
    "L_ORingTwist": "2차 O/Ring 꼬임",
    "DefectGroup": "분류"
}

valueDic = {
    "1": "b",
    "2": "t",
    "3": "ts1",
    "4": "ts2",
    "5": "ts3",
    "6": "ts4",
    "7": "bs1",
    "8": "bs2",
    "9": "bs3",
    "10": "bs4",
}
# 필드명 변환 함수
def convert_field(field_name):
    return fieldDic.get(field_name, field_name)

# 필드 값 변환 함수
def convert_value(value):
    if '#' in value:
        parts = value.split('#')
        result = valueDic.get(parts[0], parts[0]) + '#' + valueDic.get(parts[1], parts[1])
    else:
        result = valueDic.get(value, value)
    return result

def convert_result(value):
    if value == "1":
        result = "불량"
    elif value == "0":
        result = "양품"
    return result

def convert_group(value):
    if value == "1":
        result = "상단"
    elif value == "2":
        result = "하단"
    else:
        result = ''
    return result

def convert(field_name, value):
    if field_name == 'Result':
        return convert_result(value)
    elif field_name == 'DefectGroup':
        return convert_group(value)
    else:
        return convert_value(value)

# 연결 종료
cursor.close()
conn.close()

field_names = [convert_field(description[0]) for description in cursor.description]

# 새로운 Excel 파일 생성
wb = Workbook()

# 첫 번째 시트에 데이터프레임 내용 쓰기
ws1 = wb.active
ws1.title = '데이터'

ws1.append(field_names)
new_rows = []
for row in rows:
    new_row = [convert(cursor.description[idx][0], value) for idx, value in enumerate(row)]
    new_rows.append(new_row)

for row in new_rows:
    ws1.append(row)

# 통계 시트에 통계 정보 쓰기
ws2 = wb.create_sheet(title='통계')
for index, name in enumerate(fieldDic):
    if name not in ["Datetime", "Result", "DefectGroup"]:
        column = chr(ord('A') + index)
        print(column)
        ws2[f'A{index-1}'] = fieldDic[name]
        ws2[f'B{index-1}'] = f'=COUNTIF(데이터!{column}:{column}, "<>")-COUNTIF(데이터!{column}:{column}, "0")-1'

# COUNTIF 함수를 사용하여 개수 계산하는 수식 적용
ws2['D1'] = "총검사"
ws2['E1'] = "=E2 + E3"
ws2['D2'] = "양품"
ws2['E2'] = '=COUNTIF(데이터!B:B, "양품")'
ws2['F2'] = "=E2 / E1"
ws2['D3'] = "불량"
ws2['E3'] = '=COUNTIF(데이터!B:B, "불량")'
ws2['F3'] = "=E3 / E1"

ws2['D5'] = "상단"
ws2['E5'] = '=COUNTIF(데이터!Z:Z, "상단")'
ws2['F5'] = "=E5 / E3"
ws2['D6'] = "하단"
ws2['E6'] = '=COUNTIF(데이터!Z:Z, "하단")'
ws2['F6'] = "=E6 / E3"


# 퍼센트 서식
percent_style = NamedStyle(name='percent_style', number_format='0.00%')
percent_cells = ['F2', 'F3', 'F5', 'F6']
# for cell in percent_cells:
#     ws2[cell].style = percent_style

# 서식 적용
def apply_style(ws, cells_range, font_color, start_color, end_color, border_style='thin'):
    for cell in ws[cells_range]:
        for col in cell:
            col.font = Font(name='Arial', size=12, bold=False, italic=False, color=font_color)
            col.fill = PatternFill(start_color=start_color, end_color=end_color, fill_type='solid')
            col.border = Border(left=Side(border_style=border_style, color='000000'),
                                right=Side(border_style=border_style, color='000000'),
                                top=Side(border_style=border_style, color='000000'),
                                bottom=Side(border_style=border_style, color='000000'))
            col.alignment = Alignment(horizontal='center', vertical='center')

apply_style(ws2, 'A1:A23', '000000', 'FABF8F', 'FABF8F')
apply_style(ws2, 'B1:B23', '000000', 'FCD5B4', 'FCD5B4')
apply_style(ws2, 'D1:D3', '000000', 'FABF8F', 'FABF8F')
apply_style(ws2, 'E1:E3', '000000', 'FCD5B4', 'FCD5B4')
apply_style(ws2, 'D5:D6', '000000', 'FABF8F', 'FABF8F')
apply_style(ws2, 'E5:E6', '000000', 'FCD5B4', 'FCD5B4')
apply_style(ws2, 'F2:F3', '000000', 'FDE9D9', 'FDE9D9')
apply_style(ws2, 'F5:F6', '000000', 'FDE9D9', 'FDE9D9')
# apply_style(ws1, f'A2:Z{df.shape[0]+1}', '000000', 'FDE9D9', 'FABF8F')

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 10

# 데이터 시트 너비 조절
# for col in ws1.columns:
#     max_length = 0
#     column = col[0].column_letter
#     for cell in col:
#         try:
#             if len(str(cell.value)) > max_length:
#                 max_length = len(cell.value)
#         except:
#             pass
#     adjusted_width = (max_length + 2) * 1.4  # 너비를 조절하는 공식 (원하는 너비 조절 가능)
#     ws1.column_dimensions[column].width = adjusted_width

ws1.auto_filter.ref = ws1.dimensions

# 엑셀 파일 저장
wb.save('../output_file_with_stats.xlsx')