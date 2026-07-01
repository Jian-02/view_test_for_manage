import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment

class DataProcessor:
    def __init__(self, db_path):
        self.db_path = db_path
        self.fieldDic = {
            "ClientName": "생산라인",
            "Part": "부품",
            "Datetime": "검사시각",
            "Result": "검사결과",
            "H_BootTear": "1차 부트 찢어짐",
            "H_BootCurl": "1차 부트 말림",
            "H_BootBurr": "1차 부트 Burr 유/무",
            "H_BootAssembly": "1차 부트 조립 상태",
            "H_GreaseOut": "1차 목부 그리스 누유",
            "H_PBallDamage": "1차 P/Ball 찍힘 및 긁힘",
            "H_PBallLot": "1차 P/Ball 로트 타각 식별",
            "H_CRingErr": "1차 C/Ring 유/무",
            "H_CRingTwist": "1차 C/Ring 꼬임",
            "H_ORingErr": "1차 O/Ring 유/무",
            "H_ORingTwist": "1차 O/Ring 꼬임",
            "SocketDamage": "소켓 외경 찍힘 및 긁힘",
            "SocketGroove": "소켓 외경 홈 그루브 유무",
            "L_BootTear": "2차 부트 찢어짐",
            "L_BootCurl": "2차 부트 말림",
            "L_BootBurr": "2차 부트 Burr 유/무",
            "L_BootAssembly": "2차 부트 조립 상태",
            "L_GreaseOut": "2차 목부 그리스 누유",
            "L_PBallDamage": "2차 P/Ball 찍힘 및 긁힘",
            "L_CRingErr": "2차 C/Ring 유/무",
            "L_CRingTwist": "2차 C/Ring 꼬임",
            "L_ORingErr": "2차 O/Ring 유/무",
            "L_ORingTwist": "2차 O/Ring 꼬임",
            "DefectGroup": "분류"
        }
        self.valueDic = {
            "0": '',
            "1": "불량",
            "2": "t",
            "3": "ts1",
            "4": "ts2",
            "5": "ts3",
            "6": "ts4",
            "7": "bs1",
            "8": "bs2",
            "9": "bs3",
            "10": "bs4",
        }
        self.percent_style = NamedStyle(name='percent_style', number_format='0.00%')

    def connect_db(self):
        return sqlite3.connect(self.db_path, check_same_thread=False)

    def convert_field(self, field_name):
        return self.fieldDic.get(field_name)

    def convert_value(self, value):
        if '#' in value:
            parts = value.split('#')
            result = self.valueDic.get(parts[0], parts[0]) + '#' + self.valueDic.get(parts[1], parts[1])
        else:
            result = self.valueDic.get(value, value)
        return result

    def convert_result(self, value):
        if value == "1":
            result = "불량"
        elif value == "0":
            result = "양품"
        return result

    def convert_group(self, value):
        if value == "1":
            result = "상단"
        elif value == "2":
            result = "하단"
        else:
            result = '0'
        return result

    def convert(self, field_name, value):
        if field_name == 'Result':
            return self.convert_result(value)
        elif field_name == 'DefectGroup':
            return self.convert_group(value)
        else:
            return self.convert_value(value)

    def process_data_to_excel(self, part, table_name, sDate, eDate):
        current_time = datetime.now()
        conn = self.connect_db()
        cursor = conn.cursor()
        if part == '부품 전체':
            part = 'ALL'
        elif part == 'CABJ':
            part = 'CABJ'
        elif part == 'CABJ_Alpha':
            part = 'CABJ_Alpha'
        elif part == 'CABJ_Alpha_Z':
            part = 'CABJ_Alpha_Z'

        query = f"""SELECT ClientName, Part, Datetime, Result, H_BootTear, H_BootCurl, H_BootBurr, H_BootAssembly, H_GreaseOut, H_PBallDamage, H_PBallLot, H_CRingErr, H_CRingTwist, H_ORingErr, H_ORingTwist, SocketDamage, SocketGroove, L_BootTear, L_BootCurl, L_BootBurr, L_BootAssembly, L_GreaseOut, L_PBallDamage, L_CRingErr, L_CRingTwist, L_ORingErr, L_ORingTwist FROM inference_data_{table_name} WHERE Datetime >= ? AND Datetime <= ?"""

        if part.upper() != 'ALL':
            query += " AND Part = :part"
            params = (sDate, eDate, part)
        else:
            params = (sDate, eDate)

        query += " ORDER BY Datetime DESC"
        cursor.execute(query, params)
        print("test")
        rows = cursor.fetchall()
        print(f"db 조회 걸리는 시간 = {datetime.now() - current_time}")

        field_names = [self.convert_field(description[0]) for description in cursor.description]

        wb = Workbook()
        ws1 = wb.active
        ws1.title = '데이터'
        ws1.append(field_names)

        current_time = datetime.now()
        for row in rows:
            new_row = [self.convert(cursor.description[idx][0], value) for idx, value in enumerate(row)]
            ws1.append(new_row)

        print(f"엑셀 입력 걸리는 시간 = {datetime.now() - current_time}")

        ws1.auto_filter.ref = ws1.dimensions

        ws2 = wb.create_sheet(title='통계')
        # Add your statistical calculations here to ws2
        current_time = datetime.now()
        self.add_statistic(ws2)
        print(f"통계입력 걸리는 시간 = {datetime.now() - current_time}")
        current_time = datetime.now()
        self.apply_styles(ws2)
        self.adjust_column_widths(ws1)
        self.adjust_column_widths(ws2)
        print(f"스타일 적용 걸리는 시간 = {datetime.now() - current_time}")
        current_time = datetime.now()

        return wb

        cursor.close()
        conn.close()

    def add_statistic(self, ws):
        for index, name in enumerate(self.fieldDic):
            if name not in ["Datetime", "Result", "DefectGroup",  "ClientName", "Part"]:
                if index < 26:
                    column = chr(ord('A') + index)
                else:
                    column = 'A' + chr(ord('A') + index-26)
                # 인덱스가 1 이상일 때만 셀에 값을 넣음 (음수 방지)
                if index > 3:
                    ws[f'A{index-3}'] = self.fieldDic[name]
                    ws[f'B{index-3}'] = f'=COUNTIF(데이터!{column}:{column}, "<>")-COUNTIF(데이터!{column}:{column}, "0")-1'

        # 통계 추가
        ws['D1'] = "총검사"
        ws['E1'] = "=E2 + E3"
        ws['D2'] = "양품"
        ws['E2'] = '=COUNTIF(데이터!D:D, "양품")'
        ws['F2'] = "=E2 / E1"
        ws['D3'] = "불량"
        ws['E3'] = '=COUNTIF(데이터!D:D, "불량")'
        ws['F3'] = "=E3 / E1"
        # ws['D5'] = "상단"
        # ws['E5'] = '=COUNTIF(데이터!Z:Z, "상단")'
        # ws['F5'] = "=E5 / E3"
        # ws['D6'] = "하단"
        # ws['E6'] = '=COUNTIF(데이터!Z:Z, "하단")'
        # ws['F6'] = "=E6 / E3"
    def apply_styles(self, ws):
        cells_to_style = {
            'A1:A23': ('000000', 'FABF8F', 'FABF8F'),
            'B1:B23': ('000000', 'FCD5B4', 'FCD5B4'),
            'A1:A23': ('000000', 'FABF8F', 'FABF8F'),
            'B1:B23': ('000000', 'FCD5B4', 'FCD5B4'),
            'D1:D3': ('000000', 'FABF8F', 'FABF8F'),
            'E1:E3': ('000000', 'FCD5B4', 'FCD5B4'),
            # 'D5:D6': ('000000', 'FABF8F', 'FABF8F'),
            # 'E5:E6': ('000000', 'FCD5B4', 'FCD5B4'),
            'F2:F3': ('000000', 'FDE9D9', 'FDE9D9'),
            # 'F5:F6': ('000000', 'FDE9D9', 'FDE9D9'),
        }
        percent_cells = ['F2', 'F3', 'F5', 'F6']
        for cell in percent_cells:
            ws[cell].style = self.percent_style

        for cells_range, style_colors in cells_to_style.items():
            font_color, start_color, end_color = style_colors
            for cell in ws[cells_range]:
                for col in cell:
                    col.font = Font(name='Arial', size=12, bold=False, italic=False, color=font_color)
                    col.fill = PatternFill(start_color=start_color, end_color=end_color, fill_type='solid')
                    col.border = Border(left=Side(border_style='thin', color='000000'),
                                        right=Side(border_style='thin', color='000000'),
                                        top=Side(border_style='thin', color='000000'),
                                        bottom=Side(border_style='thin', color='000000'))
                    col.alignment = Alignment(horizontal='center', vertical='center')

    # Apply percent style
    def adjust_column_widths(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.4
            ws.column_dimensions[column].width = adjusted_width

